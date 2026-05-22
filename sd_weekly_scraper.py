#!/usr/bin/env python3
"""
South Dakota Campaign Finance – Weekly Committee Scraper
Runs every Monday until September 1, 2026.

Strategy:
  1. Playwright (headless Chromium) → search all committees, paginate via
     'Next Page' image link (title="Next Page"), collect cid/rid URLs
  2. requests (parallel) → fetch each committee detail page;
     decode Cloudflare email obfuscation; extract phone/email/address
  3. openpyxl → formatted Excel report with Summary sheet

Output: sd_reports/SD_Committees_YYYY-MM-DD.xlsx
"""

import datetime
import os
import re
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── config ──────────────────────────────────────────────────────────────
BASE_URL    = "https://sdcfr.sdsos.gov"
SEARCH_URL  = f"{BASE_URL}/Search/Search.aspx"
OUTPUT_DIR  = "./sd_reports"
MAX_WORKERS = 20
STOP_DATE   = datetime.date(2026, 9, 1)
TODAY       = datetime.date.today()

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
}


# ── Cloudflare email decoder ────────────────────────────────────────────

def decode_cf_email(encoded: str) -> str:
    key = int(encoded[:2], 16)
    return "".join(
        chr(int(encoded[i:i + 2], 16) ^ key)
        for i in range(2, len(encoded), 2)
    )


def extract_emails_from_soup(soup) -> list:
    """Decode all CF-obfuscated emails; exclude site contact address."""
    emails = []
    seen = set()
    for el in soup.find_all(attrs={"data-cfemail": True}):
        token = el.get("data-cfemail", "")
        if token:
            try:
                addr = decode_cf_email(token)
                if addr not in ("cfr@state.sd.us",) and addr not in seen:
                    emails.append(addr)
                    seen.add(addr)
            except Exception:
                pass
    for a in soup.find_all("a", href=re.compile(r"^mailto:", re.I)):
        addr = a["href"][7:].split("?")[0].strip()
        if addr and "@" in addr and addr not in ("cfr@state.sd.us",) and addr not in seen:
            emails.append(addr)
            seen.add(addr)
    return emails


# ── Phase 1: Playwright pagination ────────────────────────────────────────────

def scrape_all_committee_links():
    """
    Headless Chromium → submit 'All committees' search →
    paginate with 'Next Page' image button → collect cid/rid URLs.
    Returns list of dicts: {name, cid, rid, detail_url}
    """
    from playwright.sync_api import sync_playwright

    committees = []

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=True, args=["--no-sandbox"])
        page = browser.new_page(user_agent=HEADERS["User-Agent"])
        page.set_default_timeout(30_000)

        print("  Loading search page…")
        page.goto(SEARCH_URL, wait_until="domcontentloaded")
        page.click("#MainContent_rblCommittee_0")  # Search By Committee Name

        # Verify RadComboBox shows "All"
        combo = page.input_value("#ctl00_MainContent_rcbCommitteeList_Input")
        if combo.strip().lower() not in ("all", ""):
            page.fill("#ctl00_MainContent_rcbCommitteeList_Input", "All")
            time.sleep(0.5)

        print("  Submitting search…")
        page.click("#MainContent_btnSearchCommType_Name")
        page.wait_for_selector("table a[href*='cid=']", timeout=25_000)

        body = page.inner_text("body")
        m = re.search(r"(\d+) items in (\d+) pages", body)
        total_items = int(m.group(1)) if m else "?"
        total_pages = int(m.group(2)) if m else 1
        print(f"  {total_items} items across {total_pages} pages")

        def extract_links():
            """Extract committee {name, cid, rid, detail_url} from current grid page."""
            result = []
            for lnk in page.query_selector_all("table a[href*='cid=']"):
                href = lnk.get_attribute("href") or ""
                name = (lnk.inner_text() or "").strip()
                mx = re.search(r"cid=(\d+)&rid=(\d+)", href)
                if mx and name:
                    result.append({
                        "name": name,
                        "cid": mx.group(1),
                        "rid": mx.group(2),
                        "detail_url": urljoin(BASE_URL, href),
                    })
            return result

        def click_next_page():
            """Click the 'Next Page' image link (title='Next Page')."""
            el = page.query_selector('a[title="Next Page"]')
            if not el:
                return False
            first_el = page.query_selector("table a[href*='cid=']")
            first_before = first_el.inner_text().strip() if first_el else ""
            el.click()
            for _ in range(40):
                time.sleep(0.25)
                after_el = page.query_selector("table a[href*='cid=']")
                after = after_el.inner_text().strip() if after_el else ""
                if after and after != first_before:
                    return True
            return True  # proceed anyway after timeout

        # Collect page 1
        links = extract_links()
        committees.extend(links)
        print(f"    Page   1/{total_pages} — {len(committees)} committees")

        # Pages 2 … total_pages
        for pg_num in range(2, total_pages + 1):
            click_next_page()
            committees.extend(extract_links())
            if pg_num % 20 == 0 or pg_num == total_pages:
                print(f"    Page {pg_num:>3}/{total_pages} — {len(committees)} committees")

        browser.close()

    # Deduplicate by (cid, rid)
    seen, unique = set(), []
    for c in committees:
        key = (c["cid"], c["rid"])
        if key not in seen:
            seen.add(key)
            unique.append(c)
    print(f"  Unique committees: {len(unique)}")
    return unique


# ── Phase 2: Detail page parsing ─────────────────────────────────────────────

def parse_detail_page(html: str, entry: dict) -> dict:
    """
    Parse a committee detail page.

    Contact section is a 3-column layout (left | spacer | right):
      Row: Committee Chair  |   | Committee Treasurer
      Row: <chair name>     |   | <treasurer name>
      Row: Address          |   | Address
      Row: <addr>           |   | <addr>
      Row: Daytime Phone …  |   | Daytime Number
      Row: <phone>          |   | <phone>
      Row: Email            |   | Email
      Row: <CF email>       |   | <CF email>
    Indexed offsets from "Committee Chair" cell:
      +3 = chair name, +5 = treasurer name
      +9 = chair addr, +11 = treasurer addr
      +15 = chair phone, +17 = treasurer phone
    """
    soup = BeautifulSoup(html, "lxml")
    result = dict(entry)
    for k in ["chair", "treasurer", "chair_phone", "chair_email",
              "treasurer_phone", "treasurer_email", "committee_address",
              "committee_telephone", "committee_website",
              "candidate_name", "campaign_status", "committee_type"]:
        result.setdefault(k, "")

    # ── CF emails (chair first, treasurer second) ──────────────────────────────
    all_emails = extract_emails_from_soup(soup)
    if len(all_emails) >= 1:
        result["chair_email"] = all_emails[0]
    if len(all_emails) >= 2:
        result["treasurer_email"] = all_emails[1]

    # ── Flat cell list ──────────────────────────────────────────────────────────────
    all_cells = [td.get_text(" ", strip=True) for td in soup.find_all("td")]

    # ── Candidate name + status from early structured cells ─────────────────────
    for c in all_cells[:10]:
        if "Candidate Name:" in c and not result["candidate_name"]:
            m = re.search(
                r"Candidate Name\s*[:\s]+([A-Za-z ,.''\-]+?)(?:\s+Campaign Status|\s+Committee Type|$)",
                c, re.I)
            if m:
                result["candidate_name"] = m.group(1).strip()
        if "Campaign Status:" in c and not result["campaign_status"]:
            m = re.search(r"Campaign Status\s*[:\s]+(\w+)", c, re.I)
            if m:
                result["campaign_status"] = m.group(1).strip()

    # ── 3-column table: chair/treasurer at fixed offsets ──────────────────────
    chair_idx = next((i for i, c in enumerate(all_cells) if c == "Committee Chair"), None)
    if chair_idx is not None:
        def safe_cell(offset):
            idx = chair_idx + offset
            return all_cells[idx] if idx < len(all_cells) else ""

        result["chair"]           = safe_cell(3)
        result["treasurer"]       = safe_cell(5)
        result["chair_phone"]     = safe_cell(15)
        result["treasurer_phone"] = safe_cell(17)

    # ── Fallback phone extraction ────────────────────────────────────────────
    phones = [ct.strip() for ct in all_cells
              if re.match(r"^\(?\d{3}\)?[\s\-\.]\d{3}[\s\-\.]\d{4}$", ct.strip())]
    if not result["chair_phone"]     and len(phones) >= 1:
        result["chair_phone"] = phones[0]
    if not result["treasurer_phone"] and len(phones) >= 2:
        result["treasurer_phone"] = phones[1]

    # ── Committee-level fields ───────────────────────────────────────────────
    def first_value_after(label_pattern):
        for i, c in enumerate(all_cells):
            if re.fullmatch(label_pattern, c, re.I):
                for j in range(i + 1, min(i + 5, len(all_cells))):
                    v = all_cells[j].strip()
                    if v and not re.fullmatch(label_pattern, v, re.I):
                        return v
        return ""

    result["committee_address"]   = first_value_after(r"Committee Address")
    result["committee_telephone"] = first_value_after(r"Telephone Number")
    raw_web = first_value_after(r"Website")
    result["committee_website"]   = raw_web if raw_web not in ("(none)", "none") else ""

    if not result["committee_type"]:
        result["committee_type"] = first_value_after(r"Committee Type")

    return result


def fetch_detail(session: requests.Session, entry: dict, retries: int = 2) -> dict:
    for attempt in range(retries + 1):
        try:
            r = session.get(entry["detail_url"], headers=HEADERS, timeout=20)
            r.raise_for_status()
            return parse_detail_page(r.text, entry)
        except Exception as e:
            if attempt < retries:
                time.sleep(1 + attempt)
            else:
                result = dict(entry)
                result["error"] = str(e)
                return result
    return entry


def fetch_all_details(committees: list) -> list:
    session = requests.Session()
    session.headers.update(HEADERS)
    results = [None] * len(committees)
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
        fut_to_idx = {pool.submit(fetch_detail, session, c): i
                      for i, c in enumerate(committees)}
        done = 0
        for fut in as_completed(fut_to_idx):
            idx = fut_to_idx[fut]
            results[idx] = fut.result()
            done += 1
            if done % 100 == 0 or done == len(committees):
                print(f"    Fetched {done}/{len(committees)}")
    return [r for r in results if r is not None]


# ── Phase 3: Excel output ───────────────────────────────────────────────────

COLUMNS = [
    ("Committee Name",        "name"),
    ("Committee Type",        "committee_type"),
    ("Candidate",             "candidate_name"),
    ("Campaign Status",       "campaign_status"),
    ("Committee Address",     "committee_address"),
    ("Committee Telephone",   "committee_telephone"),
    ("Committee Website",     "committee_website"),
    ("Committee Chair",       "chair"),
    ("Chair Phone",           "chair_phone"),
    ("Chair Email",           "chair_email"),
    ("Committee Treasurer",   "treasurer"),
    ("Treasurer Phone",       "treasurer_phone"),
    ("Treasurer Email",       "treasurer_email"),
    ("Detail URL",            "detail_url"),
]


def save_excel(records: list, output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SD Committees"

    thin       = Side(style="thin")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill   = PatternFill("solid", fgColor="DCE6F1")
    hdr_font   = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill   = PatternFill("solid", fgColor="1F3864")
    hdr_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_align = Alignment(vertical="center", wrap_text=False)

    for ci, (header, _) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=header)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = hdr_align; cell.border = border
    ws.row_dimensions[1].height = 30

    for ri, rec in enumerate(records, 2):
        fill = alt_fill if ri % 2 == 0 else None
        for ci, (_, key) in enumerate(COLUMNS, 1):
            cell = ws.cell(row=ri, column=ci, value=rec.get(key, ""))
            cell.alignment = data_align; cell.border = border
            if fill:
                cell.fill = fill

    widths = [38, 45, 25, 12, 40, 16, 35, 25, 16, 35, 25, 16, 35, 60]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "South Dakota Campaign Finance – Committee Export"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2["A3"] = "Run Date";            ws2["B3"] = TODAY.isoformat()
    ws2["A4"] = "Total Committees";    ws2["B4"] = len(records)
    ws2["A5"] = "Source";              ws2["B5"] = SEARCH_URL
    next_run = TODAY + datetime.timedelta(days=(7 - TODAY.weekday()) % 7 or 7)
    ws2["A6"] = "Next Scheduled Run";  ws2["B6"] = next_run.isoformat()
    ws2["A8"] = "Committee Type Breakdown"; ws2["A8"].font = Font(bold=True)
    ws2["B8"] = "Count";               ws2["B8"].font = Font(bold=True)
    type_counts: dict = {}
    for r in records:
        t = r.get("committee_type") or "Unknown"
        type_counts[t] = type_counts.get(t, 0) + 1
    for i, (t, cnt) in enumerate(sorted(type_counts.items(), key=lambda x: -x[1]), 9):
        ws2.cell(row=i, column=1, value=t)
        ws2.cell(row=i, column=2, value=cnt)
    ws2.column_dimensions["A"].width = 55
    ws2.column_dimensions["B"].width = 10

    wb.save(output_path)
    print(f"  ✓ Saved: {output_path}")


# ── Main ──────────────────────────────────────────────────────────────

def main():
    print(f"\n{'='*60}")
    print(f"  SD Campaign Finance Scraper — {TODAY}")
    print(f"{'='*60}\n")

    if TODAY >= STOP_DATE:
        print(f"Stop date {STOP_DATE} reached. Exiting.")
        sys.exit(0)

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    output_path = os.path.join(OUTPUT_DIR, f"SD_Committees_{TODAY}.xlsx")

    print("[1/3] Collecting committee links…")
    committees = scrape_all_committee_links()
    if not committees:
        print("ERROR: No committees found."); sys.exit(1)

    print(f"\n[2/3] Fetching {len(committees)} detail pages…")
    detail_records = fetch_all_details(committees)
    errors = [r for r in detail_records if r.get("error")]
    if errors:
        print(f"  ⚠ {len(errors)} pages had errors")

    print(f"\n[3/3] Building Excel report…")
    final_records = []
    for d in detail_records:
        rec = {col_key: d.get(col_key, "") for _, col_key in COLUMNS}
        final_records.append(rec)
    final_records.sort(key=lambda r: (r.get("name") or "").lstrip().lower())

    save_excel(final_records, output_path)

    txt_path = os.path.join(OUTPUT_DIR, f"SD_Committees_{TODAY}_summary.txt")
    type_counts: dict = {}
    for r in final_records:
        t = r.get("committee_type") or "Unknown"
        type_counts[t] = type_counts.get(t, 0) + 1
    with open(txt_path, "w") as f:
        f.write(f"SD Campaign Finance – Committee Export\n")
        f.write(f"Run date:         {TODAY}\n")
        f.write(f"Total committees: {len(final_records)}\n")
        f.write(f"Errors:           {len(errors)}\n\n")
        f.write("Breakdown by type:\n")
        for t, cnt in sorted(type_counts.items(), key=lambda x: -x[1]):
            f.write(f"  {t}: {cnt}\n")
    print(f"  ✓ Summary: {txt_path}")
    print(f"\n✅ Done! {len(final_records)} committees → {output_path}\n")


if __name__ == "__main__":
    main()
